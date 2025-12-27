#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel工具函数库
提供Excel文件处理的通用工具函数
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlwings as xw
from datetime import datetime
import os


class ExcelUtils:
    """Excel工具类"""
    
    @staticmethod
    def read_excel_with_format(file_path):
        """
        读取Excel文件并保留格式
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            tuple: (数据框, 工作簿对象)
        """
        # 使用pandas读取数据
        df = pd.read_excel(file_path)
        
        # 使用openpyxl读取格式
        workbook = load_workbook(file_path)
        
        return df, workbook
    
    @staticmethod
    def write_excel_with_format(df, file_path, template_workbook=None):
        """
        写入Excel文件并应用格式
        
        Args:
            df: 数据框
            file_path: 输出文件路径
            template_workbook: 模板工作簿（可选）
        """
        # 使用pandas写入数据
        df.to_excel(file_path, index=False)
        
        if template_workbook:
            # 应用模板格式
            ExcelUtils.apply_format_from_template(file_path, template_workbook)
    
    @staticmethod
    def apply_format_from_template(output_file, template_workbook):
        """
        从模板应用格式到输出文件
        
        Args:
            output_file: 输出文件路径
            template_workbook: 模板工作簿
        """
        # 打开输出文件
        output_wb = load_workbook(output_file)
        output_ws = output_wb.active
        
        # 获取模板工作表的格式
        template_ws = template_workbook.active
        
        # 复制列宽
        for col in range(1, template_ws.max_column + 1):
            col_letter = get_column_letter(col)
            output_ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width
        
        # 复制行高
        for row in range(1, template_ws.max_row + 1):
            output_ws.row_dimensions[row].height = template_ws.row_dimensions[row].height
        
        # 保存修改
        output_wb.save(output_file)
    
    @staticmethod
    def detect_date_columns(df):
        """
        检测数据框中的日期列
        
        Args:
            df: 数据框
            
        Returns:
            list: 日期列名列表
        """
        date_columns = []
        
        for col in df.columns:
            # 检查列名是否包含日期相关关键词
            col_str = str(col).lower()
            if any(keyword in col_str for keyword in ['date', '时间', '日期', 'day', 'month', 'year']):
                date_columns.append(col)
                continue
            
            # 检查列数据是否包含日期值
            try:
                sample_data = df[col].dropna().head(10)
                if len(sample_data) > 0:
                    # 尝试转换为日期
                    pd.to_datetime(sample_data, errors='coerce')
                    # 如果成功转换的数量超过一半，认为是日期列
                    if len(sample_data) >= 5:
                        date_columns.append(col)
            except:
                pass
        
        return date_columns
    
    @staticmethod
    def format_excel_file(file_path):
        """
        格式化Excel文件
        
        Args:
            file_path: Excel文件路径
        """
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # 设置默认字体
        default_font = Font(name='微软雅黑', size=11)
        
        # 设置表头样式
        header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置边框
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # 应用表头样式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = default_font
                cell.border = thin_border
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 保存修改
        workbook.save(file_path)
    
    @staticmethod
    def merge_excel_files(file_list, output_file):
        """
        合并多个Excel文件
        
        Args:
            file_list: 文件路径列表
            output_file: 输出文件路径
        """
        merged_df = pd.DataFrame()
        
        for file_path in file_list:
            try:
                df = pd.read_excel(file_path)
                merged_df = pd.concat([merged_df, df], ignore_index=True)
            except Exception as e:
                print(f"读取文件 {file_path} 失败: {e}")
        
        merged_df.to_excel(output_file, index=False)
        ExcelUtils.format_excel_file(output_file)
    
    @staticmethod
    def split_excel_by_column(file_path, split_column, output_dir):
        """
        按列值拆分Excel文件
        
        Args:
            file_path: 输入文件路径
            split_column: 拆分列名
            output_dir: 输出目录
        """
        df = pd.read_excel(file_path)
        
        if split_column not in df.columns:
            raise ValueError(f"列 '{split_column}' 不存在于文件中")
        
        # 创建输出目录
        os.makedirs(output_dir, exist_ok=True)
        
        # 按列值分组
        grouped = df.groupby(split_column)
        
        for value, group_df in grouped:
            # 生成安全的文件名
            safe_value = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
            output_file = os.path.join(output_dir, f"{safe_value}.xlsx")
            
            group_df.to_excel(output_file, index=False)
            ExcelUtils.format_excel_file(output_file)


# 日期格式常量
DATE_FORMATS = [
    "%Y-%m-%d",      # 2023-10-15
    "%Y/%m/%d",      # 2023/10/15
    "%Y年%m月%d日",  # 2023年10月15日
    "%m/%d/%Y",      # 10/15/2023
    "%d/%m/%Y",      # 15/10/2023
    "%Y%m%d",        # 20231015
]


if __name__ == "__main__":
    # 测试代码
    print("Excel工具函数库测试")
    
    # 创建一个测试数据框
    test_data = {
        '日期': ['2023-10-15', '2023-10-16', '2023-10-17'],
        '姓名': ['张三', '李四', '王五'],
        '金额': [100, 200, 150]
    }
    
    df = pd.DataFrame(test_data)
    print("测试数据框:")
    print(df)
    
    # 测试日期列检测
    date_cols = ExcelUtils.detect_date_columns(df)
    print(f"检测到的日期列: {date_cols}")