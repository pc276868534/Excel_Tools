#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 日期分类工具
按日期自动分类Excel数据，支持多种格式
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import time
import threading
import xlwings as xw
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from excel_utils import ExcelUtils, DATE_FORMATS


class DateFilterTool:
    """Excel 日期分类工具类"""
    
    def __init__(self, parent):
        self.parent = parent
        self.window = parent
        # 不设置窗口标题和大小，因为现在是内嵌模式
        # self.window.title("📊 Excel 日期分类工具 ")
        # self.window.geometry("800x650")
        # self.window.resizable(True, True)
        
        # 初始化变量
        self.file_path = tk.StringVar()
        self.date_column = tk.StringVar()
        self.processing = False
        self.output_file_path = None
        self.xl_app = None
        
        # 创建主界面
        self.create_main_interface()
        
        # 内嵌模式不需要窗口居中和关闭协议
        # self.center_window()
        # self.window.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def create_main_interface(self):
        """创建主界面"""
        # 主容器
        self.main_container = tk.Frame(self.window, bg='#f5f8ff')
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 创建UI组件
        self.create_title_frame()
        self.create_file_selection_frame()
        self.create_column_selection_frame()
        self.create_options_frame()
        self.create_action_frame()
        self.create_progress_frame()
        self.create_status_frame()
    
    def create_title_frame(self):
        """创建标题区域"""
        title_frame = ExcelUtils.create_ui_frame(
            self.main_container, 
            "📊 Excel 日期分类工具 - 专业版", 
            "按日期分类 | 保留图片格式 | 统一行高"
        )
        title_frame.pack(fill=tk.X, pady=(0, 15))
    
    def create_file_selection_frame(self):
        """创建文件选择区域"""
        file_frame = ttk.LabelFrame(self.main_container, text="📁 选择Excel文件", padding=15)
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(file_frame, text="Excel文件:", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        
        entry_file = ttk.Entry(file_frame, textvariable=self.file_path, width=50)
        entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 10))
        
        ttk.Button(file_frame, text="浏览", command=self.browse_file).pack(side=tk.LEFT)
    
    def create_column_selection_frame(self):
        """创建列选择区域"""
        column_frame = ttk.LabelFrame(self.main_container, text="📅 选择日期列", padding=15)
        column_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(column_frame, text="日期列:", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        
        self.column_combo = ttk.Combobox(column_frame, textvariable=self.date_column, 
                                        font=("微软雅黑", 10), state="readonly", width=30)
        self.column_combo.pack(side=tk.LEFT, padx=(10, 0))
        
        ttk.Button(column_frame, text="刷新列", command=self.refresh_columns).pack(side=tk.LEFT, padx=(10, 0))
    
    def create_options_frame(self):
        """创建选项区域"""
        options_frame = ttk.LabelFrame(self.main_container, text="⚙️ 处理选项", padding=15)
        options_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 处理模式选择
        mode_frame = tk.Frame(options_frame)
        mode_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(mode_frame, text="处理模式:", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        
        self.processing_mode = tk.StringVar(value="fast")
        ttk.Radiobutton(mode_frame, text="快速模式 (推荐)", variable=self.processing_mode, value="fast").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Radiobutton(mode_frame, text="标准模式 (保留完整格式)", variable=self.processing_mode, value="standard").pack(side=tk.LEFT, padx=(10, 0))
        
        # 日期格式选项
        format_frame = tk.Frame(options_frame)
        format_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(format_frame, text="日期格式:", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        
        self.date_format = tk.StringVar(value="YYYY-MM-DD")
        for fmt in DATE_FORMATS:
            ttk.Radiobutton(format_frame, text=fmt, variable=self.date_format, value=fmt).pack(side=tk.LEFT, padx=(10, 0))
        
        # 是否保留原表
        self.keep_original = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="保留原工作表", variable=self.keep_original).pack(anchor=tk.W)
    
    def create_action_frame(self):
        """创建操作区域"""
        action_frame = tk.Frame(self.main_container)
        action_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.process_btn = ttk.Button(action_frame, text="🚀 开始分类处理", 
                                     command=self.start_processing, style="Accent.TButton")
        self.process_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.open_result_btn = ttk.Button(action_frame, text="📂 打开结果文件", 
                                           command=self.open_output_file,
                                           state=tk.DISABLED)
        self.open_result_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(action_frame, text="🗑️ 清空", command=self.clear_all).pack(side=tk.LEFT, padx=(0, 10))
        self.close_btn = ttk.Button(action_frame, text="❌ 返回主页", command=self.return_to_main).pack(side=tk.LEFT)
    
    def create_progress_frame(self):
        """创建进度条区域"""
        progress_frame = ttk.Frame(self.main_container)
        progress_frame.pack(fill=tk.X, pady=(5, 10))
        
        # 进度条
        self.progress = ttk.Progressbar(progress_frame, mode='determinate', length=600)
        self.progress.pack(fill=tk.X)
        
        # 进度标签
        self.progress_label = ttk.Label(progress_frame, text="就绪", font=("微软雅黑", 9))
        self.progress_label.pack()
    
    def create_status_frame(self):
        """创建状态区域"""
        status_frame = ttk.LabelFrame(self.main_container, text="📊 处理状态", padding=15)
        status_frame.pack(fill=tk.BOTH, expand=True)
        
        self.status_text = tk.Text(status_frame, height=32, font=("微软雅黑", 9), wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.status_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.status_text.yview)
        
        # 初始化状态
        self.add_status_message("🚀 日期分类工具已就绪")
        self.add_status_message("📁 请选择Excel文件开始操作...")
    
    def center_window(self):
        """窗口居中 - 内嵌模式不需要"""
        pass
        # self.window.update_idletasks()
        # width = self.window.winfo_width()
        # height = self.window.winfo_height()
        # x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        # y = (self.window.winfo_screenheight() // 2) - (height // 2)
        # self.window.geometry(f"{width}x{height}+{x}+{y}")
    
    def browse_file(self):
        """浏览文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm *.xlsb"), ("所有文件", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            self.add_status_message(f"✅ 已选择文件: {os.path.basename(filename)}")
            self.refresh_columns()
    
    def refresh_columns(self):
        """刷新列信息"""
        filename = self.file_path.get()
        
        # 使用公共工具验证文件
        is_valid, error_msg = ExcelUtils.validate_excel_file(filename)
        if not is_valid:
            self.add_status_message(f"❌ {error_msg}", is_error=True)
            return
        
        try:
            # 使用公共工具获取列名
            columns = ExcelUtils.get_excel_columns(filename)
            self.column_combo['values'] = columns
            if columns:
                self.date_column.set(columns[0])
                self.add_status_message(f"✅ 列信息刷新完成，共 {len(columns)} 列")
        except Exception as e:
            self.add_status_message(f"❌ 读取列信息失败: {str(e)}", is_error=True)
    
    def start_processing(self):
        """开始处理"""
        if self.processing:
            return
            
        if not self.validate_inputs():
            return
            
        self.processing = True
        self.process_btn.config(state=tk.DISABLED)
        self.open_result_btn.config(state=tk.DISABLED)
        self.progress["value"] = 0
        self.progress_label.config(text="准备中...")
        
        # 在后台线程中执行
        threading.Thread(target=self.process_table, daemon=True).start()
    
    def validate_inputs(self):
        """验证输入"""
        if not self.file_path.get():
            messagebox.showwarning("警告", "请选择Excel文件")
            return False
            
        if not os.path.exists(self.file_path.get()):
            messagebox.showwarning("警告", "选择的文件不存在")
            return False
            
        if not self.date_column.get():
            messagebox.showwarning("警告", "请选择日期列")
            return False
            
        return True
    
    def process_table(self):
        """表格分类处理的核心逻辑 - 根据模式选择不同的处理方式"""
        mode = self.processing_mode.get()
        
        if mode == "fast":
            self.process_table_fast_mode()
        else:
            self.process_table_standard_mode()
    
    def process_table_fast_mode(self):
        """快速模式处理 - 使用openpyxl快速处理数据，不保留图片"""
        try:
            self.add_status_message("="*50)
            self.add_status_message("🚀 使用快速模式处理...")
            start_time = time.time()
            
            # 更新进度
            self.update_progress(5, "正在准备处理...")
            
            # 使用openpyxl读取Excel文件（只读数据）
            self.update_progress(10, "📊 正在快速读取Excel文件数据...")
            wb_original = load_workbook(self.file_path.get(), data_only=True)
            ws_original = wb_original.active
            
            # 读取表头
            header = [cell.value for cell in ws_original[1]]
            if not header:
                raise ValueError("Excel文件没有表头")
            
            # 检查日期列是否存在
            date_col_name = self.date_column.get()
            if date_col_name not in header:
                raise ValueError(f"未找到日期列: {date_col_name}")
            
            date_col_idx = header.index(date_col_name) + 1
            
            # 获取数据范围
            max_row = ws_original.max_row
            
            # 读取日期数据
            self.update_progress(20, "📅 正在快速处理日期数据...")
            date_groups = {}
            total_rows = max_row - 1
            
            for row in range(2, max_row + 1):
                if not self.processing:  # 检查是否被终止
                    break
                
                date_value = ws_original.cell(row=row, column=date_col_idx).value
                
                # 使用公共工具解析日期
                date_obj = ExcelUtils.parse_date_value(date_value)
                if date_obj:
                    if date_obj not in date_groups:
                        date_groups[date_obj] = []
                    date_groups[date_obj].append(row)
                
                # 更新进度 (20-40%)
                if (row - 1) % 100 == 0 or row == max_row:
                    progress = int((row - 1) / total_rows * 20) + 20
                    self.update_progress(progress, f"正在分析日期数据: {row - 1}/{total_rows} 行")
            
            if not self.processing:
                wb_original.close()
                return
            
            invalid_count = (max_row - 1) - sum(len(rows) for rows in date_groups.values())
            if invalid_count > 0:
                self.add_status_message(f"⚠️ 发现 {invalid_count} 行无效日期数据，已跳过")
            
            # 按日期分组
            self.update_progress(45, "📂 正在按日期分类数据...")
            
            # 创建新的Excel文件
            self.update_progress(50, "💾 正在创建新文件...")
            from openpyxl import Workbook
            wb_new = Workbook()
            wb_new.remove(wb_new.active)  # 删除默认工作表
            
            # 如果需要保留原表，先复制原数据
            if self.keep_original.get():
                self.update_progress(55, "📄 正在复制原工作表...")
                ws_original_copy = wb_new.create_sheet("原数据")
                
                # 复制表头
                for col_idx, value in enumerate(header, 1):
                    ws_original_copy.cell(row=1, column=col_idx).value = value
                
                # 复制数据行
                for row in range(2, max_row + 1):
                    if not self.processing:
                        break
                    for col_idx in range(1, len(header) + 1):
                        value = ws_original.cell(row=row, column=col_idx).value
                        ws_original_copy.cell(row=row, column=col_idx).value = value
                
                # 设置原数据工作表的行高为50磅、内容居中和自动换行
                if max_row > 1:  # 有数据行
                    for row_idx in range(2, max_row + 1):
                        ws_original_copy.row_dimensions[row_idx].height = 50
                    # 设置数据区域格式
                    from openpyxl.styles import Alignment
                    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    for row in ws_original_copy.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=len(header)):
                        for cell in row:
                            cell.alignment = align_center
            
            # 为每个日期创建工作表
            total_groups = len(date_groups)
            processed = 0
            
            # 按日期排序，确保处理顺序一致
            sorted_dates = sorted(date_groups.keys())
            
            for date in sorted_dates:
                if not self.processing:
                    break
                
                row_numbers = date_groups[date]
                
                # 使用公共工具格式化工作表名称
                sheet_name = ExcelUtils.format_sheet_name(date, self.date_format.get())
                
                # 创建新工作表
                ws_new = wb_new.create_sheet(sheet_name)
                
                # 复制表头
                for col_idx, value in enumerate(header, 1):
                    ws_new.cell(row=1, column=col_idx).value = value
                
                # 复制数据行（支持换行合并）
                target_row = 2
                seen_rows = {}  # 用于记录已见的行，支持合并
                merged_count = 0  # 记录合并次数
                
                for row_num in row_numbers:
                    if not self.processing:
                        break
                    
                    # 读取当前行的数据（排除日期列）
                    row_data = []
                    for col_idx in range(1, len(header) + 1):
                        if col_idx != date_col_idx:  # 排除日期列
                            value = ws_original.cell(row=row_num, column=col_idx).value
                            row_data.append(value)
                    
                    # 创建行的唯一键（基于非日期列的数据）
                    row_key = tuple(str(v) if v is not None else '' for v in row_data)
                    
                    if row_key in seen_rows:
                        # 如果已存在相同行，执行合并逻辑
                        existing_row = seen_rows[row_key]
                        merged_count += 1  # 记录合并次数
                        
                        # 对于某些列，如果值不同，则用换行符合并
                        for col_idx in range(1, len(header) + 1):
                            if col_idx != date_col_idx:  # 排除日期列
                                new_value = ws_original.cell(row=row_num, column=col_idx).value
                                existing_value = ws_new.cell(row=existing_row, column=col_idx).value
                                
                                if new_value and new_value != existing_value:
                                    # 如果新值不为空且与现有值不同，进行换行合并
                                    if existing_value:
                                        merged_value = f"{existing_value}\n{new_value}"
                                    else:
                                        merged_value = str(new_value)
                                    ws_new.cell(row=existing_row, column=col_idx).value = merged_value
                    else:
                        # 新行，直接复制
                        for col_idx in range(1, len(header) + 1):
                            value = ws_original.cell(row=row_num, column=col_idx).value
                            ws_new.cell(row=target_row, column=col_idx).value = value
                        
                        seen_rows[row_key] = target_row
                        target_row += 1
                
                # 设置行高为50磅、内容居中和自动换行（快速模式）
                if target_row > 2:  # 有数据行
                    for row_idx in range(2, target_row):
                        ws_new.row_dimensions[row_idx].height = 50
                    # 设置数据区域格式
                    from openpyxl.styles import Alignment
                    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    for row in ws_new.iter_rows(min_row=2, max_row=target_row-1, min_col=1, max_col=len(header)):
                        for cell in row:
                            cell.alignment = align_center
                
                processed += 1
                # 更新进度 (55-90%)
                progress = int((processed / total_groups) * 35) + 55
                self.update_progress(progress, f"📊 已处理 {processed}/{total_groups} 个日期: {sheet_name}")
            
            if not self.processing:
                wb_original.close()
                return
            
            # 保存文件
            output_path = self.get_save_location()
            if not output_path:
                wb_original.close()
                return
                
            self.update_progress(95, "💾 正在保存文件...")
            wb_new.save(output_path)
            wb_original.close()
            
            # 计算耗时
            elapsed_time = time.time() - start_time
            self.update_progress(100, "✅ 快速处理完成！")
            self.add_status_message(f"✅ 快速处理完成！共创建 {len(date_groups)} 个工作表")
            self.add_status_message(f"🔄 换行合并: 共合并 {merged_count} 行重复数据")
            self.add_status_message(f"⚡ 处理耗时: {elapsed_time:.2f}秒")
            self.add_status_message(f"📁 结果文件: {os.path.basename(output_path)}")
            self.add_status_message("="*50)
            
            # 保存结果文件路径
            self.output_file_path = output_path
            
            # 显示完成消息
            self.show_completion_message(output_path, len(date_groups))
            
        except Exception as e:
            self.update_progress(0, f"❌ 快速模式处理失败: {str(e)}")
            self.add_status_message(f"❌ 快速模式处理失败: {str(e)}", is_error=True)
        finally:
            self.stop_processing()
    
    def process_table_standard_mode(self):
        """标准模式处理 - 使用xlwings保留图片和格式"""
        app = xw.App(visible=False)
        self.xl_app = app  # 保存引用以便退出时关闭
        
        try:
            self.add_status_message("="*50)
            self.add_status_message("⚙️ 使用标准模式处理...")
            start_time = time.time()
            
            self.update_progress(5, "正在准备处理...")
            
            # 获取输出路径
            output_path = self.get_save_location()
            if not output_path:
                return
            
            # 打开原始文件
            self.update_progress(10, "📊 正在读取Excel文件...")
            wb_original = app.books.open(self.file_path.get())
            ws_original = wb_original.sheets[0]
            
            # 读取表头
            header = ws_original.range('1:1').value
            if not header:
                raise ValueError("Excel文件没有表头")
            
            # 确保header是列表格式
            if not isinstance(header, list):
                header = [header]
            
            # 统计列数（用于后续操作）
            self.total_columns = len([h for h in header if h is not None])
            
            # 检查日期列是否存在
            date_col_name = self.date_column.get()
            if date_col_name not in header:
                raise ValueError(f"未找到日期列: {date_col_name}")
            
            date_col_idx = header.index(date_col_name) + 1
            
            # 获取数据范围
            last_row = ws_original.range('A' + str(ws_original.cells.last_cell.row)).end('up').row
            if last_row < 2:
                last_row = ws_original.used_range.last_cell.row
            
            # 读取日期数据（批量优化）
            self.update_progress(20, "📅 正在批量读取日期数据...")
            
            # 使用批量读取优化性能 - 一次性读取整个列
            date_col_letter = chr(64 + date_col_idx)  # 列字母 A, B, C...
            date_range = ws_original.range(f"{date_col_letter}2:{date_col_letter}{last_row}")
            date_values_raw = date_range.value
            
            # 确保返回的是列表
            if not isinstance(date_values_raw, list):
                date_values_raw = [date_values_raw]
            
            date_values = []
            valid_rows = []
            total_rows = len(date_values_raw)
            
            for i, date_value in enumerate(date_values_raw):
                if not self.processing:
                    break
                
                # 使用公共工具解析日期
                date_obj = ExcelUtils.parse_date_value(date_value)
                if date_obj:
                    date_values.append(date_obj)
                    valid_rows.append(i + 2)  # +2 因为从第2行开始
                
                # 更新进度 (20-40%)
                if (i + 1) % 100 == 0 or i == total_rows - 1:
                    progress = int((i + 1) / total_rows * 20) + 20
                    self.update_progress(progress, f"正在分析日期数据: {i + 1}/{total_rows} 行")
            
            if not self.processing:
                wb_original.close()
                app.quit()
                return
            
            invalid_count = (last_row - 1) - len(valid_rows)
            if invalid_count > 0:
                self.add_status_message(f"⚠️ 发现 {invalid_count} 行无效日期数据，已跳过")
            
            # 按日期分组（批量处理）
            self.update_progress(45, "📂 正在按日期分类数据...")
            date_groups = {}
            
            # 批量处理日期分组，减少循环开销
            for i, date in enumerate(date_values):
                if not self.processing:
                    break
                
                row_num = valid_rows[i]
                if date not in date_groups:
                    date_groups[date] = []
                date_groups[date].append(row_num)
            
            if not self.processing:
                wb_original.close()
                app.quit()
                return
            
            # 创建新的Excel文件
            self.update_progress(50, "💾 正在创建新文件...")
            wb_new = app.books.add()
            
            # 如果需要保留原表，先复制原数据
            if self.keep_original.get():
                self.update_progress(55, "📄 正在复制原工作表...")
                ws_original.copy(before=wb_new.sheets[0])
                wb_new.sheets[0].name = "原数据"
            
            # 删除默认工作表
            if len(wb_new.sheets) > 1:
                wb_new.sheets[1].delete()
            
            # 优化处理日期分组（避免COM对象线程安全问题，但提升批量性能）
            total_groups = len(date_groups)
            processed = 0
            
            # 按日期排序，确保处理顺序一致
            sorted_dates = sorted(date_groups.keys())
            
            # 安全批量创建工作表（避免COM错误）
            self.update_progress(55, "📑 正在安全创建工作表...")
            for date in sorted_dates:
                if not self.processing:
                    break
                
                try:
                    sheet_name = ExcelUtils.format_sheet_name(date, self.date_format.get())
                    ws_new = wb_new.sheets.add(sheet_name)
                    
                    # 安全复制表头
                    try:
                        ws_original.range('1:1').copy(ws_new.range('1:1'))
                    except:
                        # 如果复制失败，手动复制表头数据
                        header_values = ws_original.range('1:1').value
                        if header_values:
                            ws_new.range('1:1').value = header_values
                            
                except Exception as sheet_error:
                    self.add_status_message(f"⚠️ 创建工作表失败: {sheet_name} - {str(sheet_error)}")
                    # 跳过这个工作表，但继续处理其他工作表
                    continue
            
            # VLOOKUP风格的高效批处理（大幅提升性能）
            self.update_progress(60, "🚀 开始VLOOKUP风格批量处理...")
            
            # 一次性收集所有需要的数据行（避免重复读取）
            all_required_rows = set()
            for row_numbers in date_groups.values():
                all_required_rows.update(row_numbers)
            all_required_rows = sorted(list(all_required_rows))
            
            # 分批读取原始数据（优化安全性）
            batch_size = 50  # 进一步减少到50行，提高稳定性
            original_data_cache = {}  # {row_num: row_data}
            
            self.update_progress(65, "📖 批量读取原始数据...")
            
            for batch_start in range(0, len(all_required_rows), batch_size):
                batch_end = min(batch_start + batch_size, len(all_required_rows))
                batch_rows = all_required_rows[batch_start:batch_end]
                
                try:
                    # 方法1：尝试连续范围批量读取
                    if len(batch_rows) == 1:
                        # 单行：直接读取
                        row_num = batch_rows[0]
                        row_data = ws_original.range(f"{row_num}:{row_num}").value
                        if not isinstance(row_data, list):
                            row_data = [row_data]
                        original_data_cache[row_num] = row_data
                    else:
                        # 多行：检查是否连续
                        is_consecutive = all(batch_rows[i] == batch_rows[i-1] + 1 for i in range(1, len(batch_rows)))
                        
                        if is_consecutive:
                            # 连续行：批量读取
                            start_row = batch_rows[0]
                            end_row = batch_rows[-1]
                            range_str = f"{start_row}:{end_row}"
                            batch_data = ws_original.range(range_str).value
                            
                            if not isinstance(batch_data, list):
                                batch_data = [[batch_data]]
                            
                            # 缓存连续数据
                            for i, row_num in enumerate(batch_rows):
                                if i < len(batch_data):
                                    original_data_cache[row_num] = batch_data[i]
                                else:
                                    original_data_cache[row_num] = [None] * len(header)
                        else:
                            # 不连续行：分组处理
                            consecutive_groups = []
                            current_group = [batch_rows[0]]
                            
                            for i in range(1, len(batch_rows)):
                                if batch_rows[i] == batch_rows[i-1] + 1:
                                    current_group.append(batch_rows[i])
                                else:
                                    consecutive_groups.append(current_group)
                                    current_group = [batch_rows[i]]
                            consecutive_groups.append(current_group)
                            
                            # 批量读取每个连续组
                            for group in consecutive_groups:
                                if len(group) == 1:
                                    # 单行读取
                                    row_data = ws_original.range(f"{group[0]}:{group[0]}").value
                                    if not isinstance(row_data, list):
                                        row_data = [row_data]
                                    original_data_cache[group[0]] = row_data
                                else:
                                    # 批量读取连续组
                                    range_str = f"{group[0]}:{group[-1]}"
                                    group_data = ws_original.range(range_str).value
                                    
                                    if not isinstance(group_data, list):
                                        group_data = [[group_data]]
                                    
                                    # 缓存组数据
                                    for i, row_num in enumerate(group):
                                        if i < len(group_data):
                                            original_data_cache[row_num] = group_data[i]
                                        else:
                                            original_data_cache[row_num] = [None] * len(header)
                            
                except Exception as read_error:
                    # 方法2：回退到安全的逐行读取
                    self.add_status_message(f"⚠️ 批量读取失败，逐行读取: {str(read_error)}")
                    
                    # 检查是否是COM错误，如果是则暂停一下
                    if "COM" in str(read_error) or "-2147352567" in str(read_error):
                        time.sleep(0.1)  # 短暂暂停让Excel恢复
                    
                    for row_num in batch_rows:
                        try:
                            row_data = ws_original.range(f"{row_num}:{row_num}").value
                            if not isinstance(row_data, list):
                                row_data = [row_data]
                            original_data_cache[row_num] = row_data
                        except Exception as single_error:
                            # 如果单行读取也失败，使用空数据
                            original_data_cache[row_num] = [None] * len(header)
                            # 如果多个单行都失败，可能是Excel状态问题
                            if "COM" in str(single_error) or "-2147352567" in str(single_error):
                                time.sleep(0.05)  # 更短的暂停
                
                # 更新进度（更频繁）
                progress = int((batch_end / len(all_required_rows)) * 10) + 65
                self.update_progress(progress, f"📖 读取数据: {batch_end}/{len(all_required_rows)} 行 (缓存{len(original_data_cache)}行)")
            
            # 批量处理所有工作表（内存中处理）
            self.update_progress(75, "💾 批量处理工作表数据...")
            
            for i, date in enumerate(sorted_dates):
                if not self.processing:
                    break
                
                row_numbers = date_groups[date]
                sheet_name = ExcelUtils.format_sheet_name(date, self.date_format.get())
                ws_new = wb_new.sheets[sheet_name]
                
                # 在内存中准备工作表数据
                if row_numbers:
                    # 构建完整的工作表数据（包含表头）
                    table_data = []
                    table_data.append(header)  # 添加表头
                    
                    # 从缓存中批量获取数据行
                    for row_num in row_numbers:
                        if row_num in original_data_cache:
                            table_data.append(original_data_cache[row_num])
                        else:
                            table_data.append([None] * len(header))
                    
                    # 一次性写入整个工作表（vlookup核心优化）
                    try:
                        # 写入所有数据到工作表
                        target_range = ws_new.range("A1").resize(len(table_data), len(header))
                        target_range.value = table_data
                        target_row = len(table_data) + 1
                        
                        # 批量设置格式（vlookup方式）- 增强稳定性版本
                        if len(table_data) > 1:  # 有数据行时才设置格式
                            format_success = False
                            
                            # 方法1：尝试分步设置格式
                            try:
                                # 暂停一下让写入操作完成
                                time.sleep(0.05)
                                
                                # 设置数据区域格式 - 分步设置
                                data_range = ws_new.range(f"A2:{chr(64 + len(header))}{len(table_data)}")
                                
                                # 分别设置各个属性，避免COM冲突
                                try:
                                    data_range.api.WrapText = True
                                except:
                                    pass
                                    
                                try:
                                    time.sleep(0.02)
                                    data_range.api.VerticalAlignment = -4108
                                except:
                                    pass
                                    
                                try:
                                    time.sleep(0.02)
                                    data_range.api.HorizontalAlignment = -4108
                                except:
                                    pass
                                
                                # 行高设置 - 采用更保守的策略
                                try:
                                    time.sleep(0.02)
                                    if len(table_data) > 200:
                                        # 超大数据量：分中批次设置，每批次20行
                                        for batch_start in range(2, len(table_data) + 1, 20):
                                            batch_end = min(batch_start + 19, len(table_data))
                                            try:
                                                ws_new.range(f"{batch_start}:{batch_end}").api.RowHeight = 50
                                                time.sleep(0.01)  # 每批次后暂停
                                            except:
                                                # 中批次失败，对这个批次逐行设置
                                                for row in range(batch_start, batch_end + 1):
                                                    try:
                                                        ws_new.cells(row, 1).api.RowHeight = 50
                                                        time.sleep(0.005)  # 每行后短暂暂停
                                                    except:
                                                        pass
                                    elif len(table_data) > 50:
                                        # 大数据量：分小批次设置，每批次10行
                                        for batch_start in range(2, len(table_data) + 1, 10):
                                            batch_end = min(batch_start + 9, len(table_data))
                                            try:
                                                ws_new.range(f"{batch_start}:{batch_end}").api.RowHeight = 50
                                                time.sleep(0.01)
                                            except:
                                                # 小批次也失败，逐行设置这个批次
                                                for row in range(batch_start, batch_end + 1):
                                                    try:
                                                        ws_new.cells(row, 1).api.RowHeight = 50
                                                        time.sleep(0.005)
                                                    except:
                                                        pass
                                    else:
                                        # 小数据量：逐行设置，但加上延迟
                                        for row in range(2, len(table_data) + 1):
                                            try:
                                                ws_new.cells(row, 1).api.RowHeight = 50
                                                time.sleep(0.003)  # 减少延迟
                                            except:
                                                pass
                                    
                                    format_success = True
                                except Exception as row_height_error:
                                    self.add_status_message(f"⚠️ 行高设置失败: {str(row_height_error)}")
                                    format_success = False
                                
                            except Exception as format_error:
                                self.add_status_message(f"⚠️ 分步格式设置失败，跳过: {str(format_error)}")
                            
                            # 如果所有格式设置都失败，至少保证数据写入成功
                            if not format_success:
                                self.add_status_message(f"ℹ️ 格式设置跳过，数据写入成功")
                        
                    except Exception as write_error:
                        # 批量写入失败，回退到逐行写入
                        self.add_status_message(f"⚠️ 批量写入失败，逐行写入: {str(write_error)}")
                        for row_idx in range(1, len(table_data)):  # 跳过表头
                            try:
                                ws_new.range(f"{row_idx + 1}:{row_idx + 1}").value = table_data[row_idx]
                            except:
                                for col, value in enumerate(table_data[row_idx], 1):
                                    try:
                                        ws_new.cells(row_idx + 1, col).value = value
                                    except:
                                        pass
                        target_row = len(table_data)
                        
                else:
                    target_row = 2
                
                if not self.processing:
                    break
                
                processed += 1
                # 更新进度 (75-90%)
                progress = int((processed / total_groups) * 15) + 75
                self.update_progress(progress, f"📊 已处理 {processed}/{total_groups} 个日期: {sheet_name}")
            
            if not self.processing:
                wb_new.close()
                wb_original.close()
                app.quit()
                return
            
            # 保存文件
            self.update_progress(95, "💾 正在保存文件...")
            wb_new.save(output_path)
            wb_new.close()
            wb_original.close()
            
            # 计算耗时
            elapsed_time = time.time() - start_time
            self.update_progress(100, "✅ 处理完成！")
            self.add_status_message(f"✅ 处理完成！共创建 {len(date_groups)} 个工作表")
            self.add_status_message(f"⏱️ 处理耗时: {elapsed_time:.2f}秒")
            self.add_status_message(f"📁 结果文件: {os.path.basename(output_path)}")
            self.add_status_message("="*50)
            
            # 保存结果文件路径
            self.output_file_path = output_path
            
            # 显示完成消息
            self.show_completion_message(output_path, len(date_groups))
            
        except Exception as e:
            self.update_progress(0, f"❌ 标准模式处理失败: {str(e)}")
            self.add_status_message(f"❌ 标准模式处理失败: {str(e)}", is_error=True)
        finally:
            app.quit()
            self.stop_processing()
    
    def set_row_height(self, worksheet, start_row, end_row, height):
        """设置行高（兼容版本）"""
        try:
            # 逐行设置行高（兼容模式）
            for row in range(start_row, end_row + 1):
                worksheet.cells(row, 1).api.RowHeight = height
        except Exception as e:
            self.add_status_message(f"⚠️ 行高设置部分失败: {str(e)}")
    
    def set_row_height_batch(self, worksheet, start_row, end_row, height):
        """批量设置行高（安全优化版本）"""
        try:
            # 方法1：小批量设置（更安全）
            batch_size = 50  # 每批50行
            total_rows = end_row - start_row + 1
            
            for batch_start in range(0, total_rows, batch_size):
                batch_end = min(batch_start + batch_size, total_rows)
                actual_start = start_row + batch_start
                actual_end = start_row + batch_end - 1
                
                try:
                    # 小批量设置
                    range_str = f"{actual_start}:{actual_end}"
                    worksheet.range(range_str).api.RowHeight = height
                except:
                    # 如果小批量失败，使用更小批量
                    mini_batch_size = 10
                    for mini_start in range(actual_start, actual_end + 1, mini_batch_size):
                        mini_end = min(mini_start + mini_batch_size - 1, actual_end)
                        try:
                            mini_range = f"{mini_start}:{mini_end}"
                            worksheet.range(mini_range).api.RowHeight = height
                        except:
                            # 最后兜底：逐行设置
                            for row in range(mini_start, mini_end + 1):
                                try:
                                    worksheet.cells(row, 1).api.RowHeight = height
                                except:
                                    pass  # 忽略个别行设置失败
            
            self.add_status_message(f"✅ 安全批量设置行高完成: {start_row}-{end_row} -> {height}磅")
            
        except Exception as e:
            # 兜底方案：使用原始逐行设置方法
            self.add_status_message(f"⚠️ 批量设置完全失败，使用逐行设置: {str(e)}")
            self.set_row_height(worksheet, start_row, end_row, height)
    
    def set_row_height_vlookup_style(self, worksheet, start_row, end_row, height):
        """VLOOKUP风格的行高设置（更高性能）"""
        try:
            # VLOOKUP方式：根据数据量选择最优策略
            if end_row >= start_row:
                total_rows = end_row - start_row + 1
                
                if total_rows > 100:
                    # 大数据量：一次性批量设置
                    try:
                        range_str = f"{start_row}:{end_row}"
                        worksheet.range(range_str).api.RowHeight = height
                        self.add_status_message(f"✅ VLOOKUP批量设置行高: {total_rows}行")
                    except:
                        # 备选方案
                        self.set_row_height_batch(worksheet, start_row, end_row, height)
                else:
                    # 小数据量：逐行设置（更精确）
                    for row in range(start_row, end_row + 1):
                        try:
                            worksheet.cells(row, 1).api.RowHeight = height
                        except:
                            pass  # 忽略个别失败
                    self.add_status_message(f"✅ VLOOKUP逐行设置行高: {total_rows}行")
                    
        except Exception as e:
            # 兜底方案
            self.set_row_height_batch(worksheet, start_row, end_row, height)
    
    def get_save_location(self):
        """获取保存位置"""
        default_name = f"分类表格_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = ExcelUtils.get_save_location(default_name, "保存分类结果文件")
        
        if not output_path:
            self.add_status_message("操作已取消")
            return None
            
        return output_path
    
    def show_completion_message(self, output_path, sheet_count):
        """显示完成消息"""
        mode = self.processing_mode.get()
        mode_text = "快速模式" if mode == "fast" else "标准模式（保留格式）"
        
        result_msg = f"""✅ 表格分类处理完成！

📊 处理结果：
• 共创建 {sheet_count} 个工作表
• 处理模式：{mode_text}
• 结果文件已保存到：{output_path}

💡 功能说明：
• 按日期自动分类数据
• {'快速模式：仅处理数据，性能最优' if mode == 'fast' else '标准模式：保留图片和格式，处理较慢'}
        """
        
        self.window.after(0, lambda: messagebox.showinfo("完成", result_msg))
        self.window.after(0, lambda: self.open_result_btn.config(state=tk.NORMAL))
    
    def add_status_message(self, msg, is_error=False):
        """添加状态消息"""
        def update():
            self.status_text.insert(tk.END, f"{msg}\n")
            if is_error:
                self.status_text.tag_add("error", "end-2l", "end-1l")
                self.status_text.tag_config("error", foreground="red")
            self.status_text.see(tk.END)
        
        self.window.after(0, update)
    
    def update_progress(self, value, message):
        """更新进度条"""
        def update():
            self.progress["value"] = value
            self.progress_label.config(text=message)
            self.add_status_message(message)
        self.window.after(0, update)
    
    def clear_all(self):
        """清空所有选择"""
        self.file_path.set("")
        self.date_column.set("")
        self.column_combo['values'] = []
        self.output_file_path = None
        self.open_result_btn.config(state=tk.DISABLED)
        
        self.status_text.delete(1.0, tk.END)
        self.add_status_message("✅ 已清空所有选择，请重新选择文件...")
    
    def open_output_file(self):
        """打开结果文件"""
        if self.output_file_path and os.path.exists(self.output_file_path):
            try:
                import sys
                import subprocess
                if sys.platform == 'win32':
                    os.startfile(self.output_file_path)
                elif sys.platform == 'darwin':  # macOS
                    subprocess.call(['open', self.output_file_path])
                else:  # linux
                    subprocess.call(['xdg-open', self.output_file_path])
                self.add_status_message(f"📂 正在打开结果文件: {os.path.basename(self.output_file_path)}")
            except Exception as e:
                self.add_status_message(f"❌ 打开文件失败: {str(e)}", is_error=True)
        else:
            self.add_status_message("❌ 找不到结果文件，请先执行分类处理", is_error=True)
    
    def stop_processing(self):
        """停止处理"""
        self.processing = False
        
        # 关闭xlwings应用
        if self.xl_app:
            try:
                self.xl_app.quit()
            except:
                pass
            self.xl_app = None
        
        # 恢复按钮状态
        self.window.after(0, lambda: self.process_btn.config(state=tk.NORMAL))
        
        # 如果结果文件存在，启用打开按钮
        if self.output_file_path and os.path.exists(self.output_file_path):
            self.window.after(0, lambda: self.open_result_btn.config(state=tk.NORMAL))
        
        self.add_status_message("❌ 处理已被终止")
    
    def return_to_main(self):
        """返回主页"""
        # 检查处理状态
        if self.processing:
            if messagebox.askokcancel("停止处理", "表格分类处理正在进行中，确定要停止并返回主页吗？"):
                self.stop_processing()
                # 调用父级的show_home_page方法
                self.parent.master.show_home_page()
            return
        else:
            # 直接返回主页
            self.parent.master.show_home_page()
    
    def on_close(self):
        """窗口关闭事件处理 - 内嵌模式不需要"""
        # 检查处理状态，但不关闭窗口（内嵌模式由主窗口管理）
        if self.processing:
            if messagebox.askokcancel("停止处理", "表格分类处理正在进行中，确定要停止吗？"):
                self.stop_processing()
        # 内嵌模式下不关闭窗口，由主窗口管理
        # self.window.destroy()
