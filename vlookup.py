#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel VLOOKUP 工具
强大的Excel数据查找和匹配工具
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import time
import threading
import queue
import xlwings as xw
import sys
import subprocess
import openpyxl
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from excel_utils import ExcelUtils, DATE_FORMATS


class VlookupTool:
    """Excel VLOOKUP工具类"""
    
    def __init__(self, parent):
        self.parent = parent
        self.window = parent
        # 不设置窗口标题和大小，因为现在是内嵌模式
        # self.window.title("🔍 Excel VLOOKUP 工具 - 专业版")
        # self.window.geometry("900x700")
        # self.window.resizable(True, True)
        
        # 初始化变量
        self.setup_variables()
        # 创建主界面
        self.create_main_interface()
        # 设置消息队列
        self.setup_message_queue()
        
        # 内嵌模式不需要窗口居中和关闭协议
        # self.center_window()
        # self.window.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def setup_variables(self):
        """初始化变量"""
        self.file_a_path = tk.StringVar()
        self.file_b_path = tk.StringVar()
        self.output_file_path = None
        self.processing = False
        self.xl_app = None
        self.selected_column = tk.StringVar(value="追加到最后一列")
        self.message_queue = queue.Queue()
        self.result_column = tk.StringVar(value="")
        self.not_found_value = "-"  # 找不到的值用"-"代替
        self.batch_size = 500  # 批量处理的行数
        self.thread_count = 4  # 并行处理线程数
    
    def setup_message_queue(self):
        """设置消息队列处理"""
        self.window.after(100, self.process_queue)
    
    def create_main_interface(self):
        """创建主界面"""
        # 主容器
        self.main_container = tk.Frame(self.window, bg='#f5f8ff')
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 创建UI组件
        self.create_title_frame()
        self.create_file_selection_frame()
        self.create_column_selection_frame()
        self.create_action_frame()
        self.create_progress_frame()
        self.create_status_frame()
    
    def create_title_frame(self):
        """创建标题区域"""
        title_frame = ExcelUtils.create_ui_frame(
            self.main_container, 
            "🔍 Excel VLOOKUP 工具", 
            "支持批量查询 | 多线程处理 | 查找指定数据"
        )
        title_frame.pack(fill=tk.X, pady=(0, 15))
    
    def create_file_selection_frame(self):
        """创建文件选择区域"""
        file_frame = ttk.LabelFrame(self.main_container, text="📁 文件选择", padding=15)
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 主表文件
        file_a_frame = tk.Frame(file_frame, bg='white')
        file_a_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(file_a_frame, text="主表文件:", font=("微软雅黑", 10), 
                bg='white').pack(side=tk.LEFT, padx=(0, 10))
        
        entry_a = ttk.Entry(file_a_frame, textvariable=self.file_a_path, 
                           state="readonly", width=50)
        entry_a.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(file_a_frame, text="浏览", command=self.browse_file_a,
                 style="Accent.TButton").pack(side=tk.LEFT)
        
        # 参考表文件
        file_b_frame = tk.Frame(file_frame, bg='white')
        file_b_frame.pack(fill=tk.X)
        
        tk.Label(file_b_frame, text="参考表文件:", font=("微软雅黑", 10), 
                bg='white').pack(side=tk.LEFT, padx=(0, 10))
        
        entry_b = ttk.Entry(file_b_frame, textvariable=self.file_b_path, 
                           state="readonly", width=50)
        entry_b.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(file_b_frame, text="浏览", command=self.browse_file_b,
                 style="Accent.TButton").pack(side=tk.LEFT)
    
    def create_column_selection_frame(self):
        """创建列选择区域"""
        column_frame = ttk.LabelFrame(self.main_container, text="🔧 列设置", padding=15)
        column_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 创建网格布局
        column_frame.columnconfigure(1, weight=1)
        column_frame.columnconfigure(3, weight=1)
        
        # 主表查找列
        tk.Label(column_frame, text="主表查找列:", font=("微软雅黑", 10), 
                bg='white').grid(row=0, column=0, sticky=tk.W, pady=5)
        
        self.column_a_combo = ttk.Combobox(column_frame, font=("微软雅黑", 10), state="readonly")
        self.column_a_combo.grid(row=0, column=1, sticky=tk.EW, padx=(5, 20), pady=5)
        
        # 参考表搜索列
        tk.Label(column_frame, text="参考表搜索列:", font=("微软雅黑", 10), 
                bg='white').grid(row=0, column=2, sticky=tk.W, pady=5)
        
        self.column_b_combo = ttk.Combobox(column_frame, font=("微软雅黑", 10), state="readonly")
        self.column_b_combo.grid(row=0, column=3, sticky=tk.EW, pady=5)
        
        # 参考表结果列
        tk.Label(column_frame, text="参考表结果列:", font=("微软雅黑", 10), 
                bg='white').grid(row=1, column=0, sticky=tk.W, pady=5)
        
        self.column_result_combo = ttk.Combobox(column_frame, font=("微软雅黑", 10), state="readonly")
        self.column_result_combo.grid(row=1, column=1, sticky=tk.EW, padx=(5, 20), pady=5)
        
        # 结果插入列
        tk.Label(column_frame, text="结果插入到列:", font=("微软雅黑", 10), 
                bg='white').grid(row=1, column=2, sticky=tk.W, pady=5)
        
        self.result_column_combo = ttk.Combobox(column_frame, textvariable=self.result_column, 
                                              font=("微软雅黑", 10), state="readonly")
        self.result_column_combo.grid(row=1, column=3, sticky=tk.EW, pady=5)
        
        # 刷新按钮
        refresh_btn = ttk.Button(column_frame, text="🔄 刷新列信息", command=self.refresh_columns,
                               style="Accent.TButton")
        refresh_btn.grid(row=2, column=0, columnspan=4, pady=(5, 0))
    
    def create_action_frame(self):
        """创建操作区域"""
        action_frame = tk.Frame(self.main_container)
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 执行按钮
        self.execute_btn = ttk.Button(action_frame, text="🚀 执行VLOOKUP", 
                                     command=self.start_vlookup_process,
                                     style="Accent.TButton")
        self.execute_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 打开结果按钮
        self.open_result_btn = ttk.Button(action_frame, text="📂 打开结果文件", 
                                        command=self.open_output_file,
                                        state=tk.DISABLED)
        self.open_result_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 清空按钮
        ttk.Button(action_frame, text="🗑️ 清空", command=self.clear_all).pack(side=tk.LEFT, padx=(0, 10))
        
        # 退出按钮
        self.close_btn = ttk.Button(action_frame, text="❌ 返回主页", command=self.return_to_main).pack(side=tk.LEFT)
    
    def create_progress_frame(self):
        """创建进度条区域"""
        progress_frame = ttk.Frame(self.main_container)
        progress_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 进度条
        self.progress = ttk.Progressbar(progress_frame, mode='determinate', length=600)
        self.progress.pack(fill=tk.X)
        
        # 进度标签
        self.progress_label = ttk.Label(progress_frame, text="就绪")
        self.progress_label.pack()
    
    def create_status_frame(self):
        """创建状态区域"""
        status_frame = ttk.LabelFrame(self.main_container, text="📊 处理状态", padding=15)
        status_frame.pack(fill=tk.BOTH, expand=True)
        
        # 状态文本框
        self.status_text = tk.Text(status_frame, height=20, font=("微软雅黑", 9), 
                                 wrap=tk.WORD, state=tk.DISABLED)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.status_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.status_text.yview)
        
        # 初始化状态
        self.add_message("🚀 Excel VLOOKUP 工具已就绪")
        self.add_message("📁 请选择文件开始操作...")
    
    def center_window(self):
        """窗口居中 - 内嵌模式不需要"""
        pass
        # self.window.update_idletasks()
        # width = self.window.winfo_width()
        # height = self.window.winfo_height()
        # x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        # y = (self.window.winfo_screenheight() // 2) - (height // 2)
        # self.window.geometry(f"{width}x{height}+{x}+{y}")
    
    def browse_file_a(self):
        """浏览主表文件"""
        filename = filedialog.askopenfilename(
            title="选择主表文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm *.xlsb"), ("所有文件", "*.*")]
        )
        if filename:
            self.file_a_path.set(filename)
            self.add_message(f"✅ 已选择主表文件: {os.path.basename(filename)}")
            self.refresh_columns()
    
    def browse_file_b(self):
        """浏览参考表文件"""
        filename = filedialog.askopenfilename(
            title="选择参考表文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm *.xlsb"), ("所有文件", "*.*")]
        )
        if filename:
            self.file_b_path.set(filename)
            self.add_message(f"✅ 已选择参考表文件: {os.path.basename(filename)}")
            self.refresh_columns()
    
    def refresh_columns(self):
        """刷新列信息 - 使用pandas快速读取"""
        file_a = self.file_a_path.get()
        file_b = self.file_b_path.get()
        
        if not file_a and not file_b:
            self.add_message("请先选择Excel文件", is_error=True)
            return
        
        def refresh_task():
            try:
                if file_a and os.path.exists(file_a):
                    try:
                        self.add_message(f"正在读取主表列信息...")
                        df_a = pd.read_excel(file_a, nrows=0)
                        columns_a = list(df_a.columns)
                        
                        self.window.after(0, lambda: self.update_combo_values(
                            self.column_a_combo, columns_a, "主表"
                        ))
                        
                        # 更新结果列选择
                        self.window.after(0, lambda: self.update_combo_values(
                            self.result_column_combo, columns_a, "主表结果列"
                        ))
                        
                    except Exception as e:
                        self.window.after(0, lambda: self.add_message(f"❌ 读取主表列名失败: {str(e)}", is_error=True))
                
                if file_b and os.path.exists(file_b):
                    try:
                        self.add_message(f"正在读取参考表列信息...")
                        df_b = pd.read_excel(file_b, nrows=0)
                        columns_b = list(df_b.columns)
                        
                        self.window.after(0, lambda: self.update_combo_values(
                            self.column_b_combo, columns_b, "参考表搜索"
                        ))
                        
                        self.window.after(0, lambda: self.update_combo_values(
                            self.column_result_combo, columns_b, "参考表结果"
                        ))
                        
                    except Exception as e:
                        self.window.after(0, lambda: self.add_message(f"❌ 读取参考表列名失败: {str(e)}", is_error=True))
                
                self.window.after(0, lambda: self.add_message("✅ 列信息刷新完成"))
                
            except Exception as e:
                self.window.after(0, lambda: self.add_message(f"❌ 刷新列信息失败: {str(e)}", is_error=True))
        
        # 在后台线程中执行
        thread = threading.Thread(target=refresh_task, daemon=True)
        thread.start()
    
    def update_combo_values(self, combo, values, source_name):
        """更新下拉框的值"""
        combo['values'] = values
        if values:
            combo.set(values[0])
            self.add_message(f"  {source_name}列: {len(values)} 列")
    
    def start_vlookup_process(self):
        """启动VLOOKUP处理流程"""
        if self.processing:
            return
            
        self.processing = True
        self.execute_btn.config(state=tk.DISABLED)
        self.progress["value"] = 0
        self.progress_label.config(text="准备中...")
        
        # 在后台线程中执行VLOOKUP
        threading.Thread(target=self.execute_vlookup, daemon=True).start()
    
    def execute_vlookup(self):
        """执行VLOOKUP的核心逻辑"""
        try:
            # 1. 检查文件是否存在
            if not self.validate_files():
                return
                
            # 2. 获取保存位置
            output_path = self.get_save_location()
            if not output_path:
                return
                
            # 3. 使用优化的混合方案处理Excel（openpyxl + xlwings）
            # 用户可选择处理模式
            if messagebox.askyesno("处理模式选择", "是否使用快速模式？\n\n快速模式：使用openpyxl进行数据查找，大幅提高处理速度\n标准模式：保持原有逻辑，保证格式完整性\n\n推荐使用快速模式，除非您需要保留非常复杂的格式"):
                self.process_with_hybrid_mode(output_path)
            else:
                self.process_with_xlwings(output_path)
            
            # 4. 显示完成消息
            self.show_completion_message()
            
        except Exception as e:
            self.add_message(f"❌ 错误: {str(e)}", is_error=True)
        finally:
            self.processing = False
            self.window.after(0, lambda: self.execute_btn.config(state=tk.NORMAL))
            self.window.after(0, lambda: self.progress_label.config(text="完成"))
    
    def validate_files(self):
        """验证输入文件"""
        file_a = self.file_a_path.get()
        file_b = self.file_b_path.get()
        
        if not file_a or not file_b:
            self.window.after(0, lambda: messagebox.showwarning("警告", "请选择两个Excel文件"))
            return False
            
        if not os.path.exists(file_a) or not os.path.exists(file_b):
            self.window.after(0, lambda: messagebox.showwarning("警告", "选择的文件不存在"))
            return False
            
        if not all([self.column_a_combo.get(), self.column_b_combo.get(), self.column_result_combo.get()]):
            self.window.after(0, lambda: messagebox.showwarning("警告", "请选择所有必需的列"))
            return False
            
        return True
    
    def get_save_location(self):
        """获取保存位置"""
        default_name = f"VLOOKUP_结果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = ExcelUtils.get_save_location(default_name, "保存结果文件")
        
        if not output_path:
            self.add_message("操作已取消")
            return None
            
        return output_path
    
    def process_with_xlwings(self, output_path):
        """使用xlwings处理Excel"""
        start_time = time.time()
        self.add_message("="*50)
        self.add_message("开始执行VLOOKUP操作...")
        
        app = xw.App(visible=False)
        self.xl_app = app  # 保存引用以便关闭
        
        try:
            # 1. 读取参考表数据 (20%)
            self.update_progress(20, "正在读取参考表数据...")
            wb_ref = app.books.open(self.file_b_path.get())
            lookup_dict = self.build_lookup_dict(wb_ref)
            wb_ref.close()
            
            # 2. 处理主表 (50%)
            self.update_progress(50, "正在处理主表...")
            wb_main = app.books.open(self.file_a_path.get())
            self.process_main_table_fast(wb_main, lookup_dict)
            
            # 3. 保存结果 (100%)
            self.update_progress(100, f"正在保存到: {os.path.basename(output_path)}")
            wb_main.save(output_path)
            wb_main.close()
            
            self.output_file_path = output_path
            
            # 4. 计算耗时
            elapsed_time = time.time() - start_time
            self.add_message(f"✅ 处理完成！耗时 {elapsed_time:.2f}秒")
            self.add_message(f"📁 结果文件: {os.path.basename(output_path)}")
            self.add_message("="*50)
            
        except Exception as e:
            self.update_progress(0, f"处理失败: {str(e)}")
            raise e
        finally:
            app.quit()
            self.xl_app = None
    
    def process_with_hybrid_mode(self, output_path):
        """使用混合模式处理Excel（openpyxl数据查找 + xlwings格式设置）"""
        start_time = time.time()
        self.add_message("="*50)
        self.add_message("🚀 开始执行快速VLOOKUP操作（混合模式）...")
        
        try:
            # 1. 使用openpyxl快速构建查找字典 (10%)
            self.update_progress(10, "正在快速读取参考表数据...")
            lookup_dict = self.build_lookup_dict_with_openpyxl()
            
            # 2. 使用openpyxl快速处理主表数据 (40%)
            self.update_progress(40, "正在快速处理主表数据...")
            temp_data_file = self.process_main_table_with_openpyxl(lookup_dict)
            
            # 3. 使用xlwings设置格式 (70%)
            self.update_progress(70, "正在设置Excel格式...")
            self.format_with_xlwings(temp_data_file, output_path)
            
            # 4. 清理临时文件
            if os.path.exists(temp_data_file):
                os.remove(temp_data_file)
            
            self.output_file_path = output_path
            
            # 5. 计算耗时 (100%)
            elapsed_time = time.time() - start_time
            self.update_progress(100, f"快速处理完成！耗时 {elapsed_time:.2f}秒")
            self.add_message(f"⚡ 快速处理完成！耗时 {elapsed_time:.2f}秒")
            self.add_message(f"📁 结果文件: {os.path.basename(output_path)}")
            self.add_message("="*50)
            
        except Exception as e:
            self.update_progress(0, f"快速处理失败: {str(e)}")
            raise e
    
    def build_lookup_dict(self, wb_ref):
        """构建查找字典 - 单线程版本（避免COM对象线程安全问题）"""
        ws_ref = wb_ref.sheets[0]
        
        # 获取列索引
        col_b_search = self.column_b_combo.get()
        col_b_result = self.column_result_combo.get()
        
        # 读取表头
        header = ws_ref.range('1:1').value
        if not header:
            raise ValueError("参考表没有表头")
        
        try:
            search_col_idx = header.index(col_b_search) + 1
            result_col_idx = header.index(col_b_result) + 1
        except ValueError as e:
            raise ValueError(f"参考表中未找到指定列: {e}")
        
        # 获取数据范围
        last_row = ws_ref.range('A' + str(ws_ref.cells.last_cell.row)).end('up').row
        if last_row < 2:
            last_row = ws_ref.used_range.last_cell.row
        
        # 单线程处理（避免COM对象线程安全问题）
        lookup_dict = {}
        
        # 一次性读取所有数据，避免多线程冲突
        search_range = ws_ref.range(f"{chr(64 + search_col_idx)}2:{chr(64 + search_col_idx)}{last_row}")
        result_range = ws_ref.range(f"{chr(64 + result_col_idx)}2:{chr(64 + result_col_idx)}{last_row}")
        
        search_values = search_range.value
        result_values = result_range.value
        
        # 确保返回的是列表
        if not isinstance(search_values, list):
            search_values = [search_values]
        if not isinstance(result_values, list):
            result_values = [result_values]
        
        # 单线程处理数据
        for i in range(len(search_values)):
            search_value = search_values[i]
            result_value = result_values[i] if i < len(result_values) else None
            
            if search_value is not None:
                key = str(search_value).strip()
                if result_value is not None:
                    lookup_dict[key] = str(result_value).strip()
                else:
                    lookup_dict[key] = ""
        
        self.add_message(f"✅ 参考表数据加载完成，共 {len(lookup_dict)} 条记录")
        return lookup_dict
    
    def process_main_table_fast(self, wb_main, lookup_dict):
        """处理主表数据 - 单线程版本（避免COM对象线程安全问题）"""
        ws_main = wb_main.sheets[0]
        
        # 获取列索引
        col_a_lookup = self.column_a_combo.get()
        
        # 读取表头
        header_main = ws_main.range('1:1').value
        if not header_main:
            raise ValueError("主表没有表头")
        
        try:
            lookup_col_idx = header_main.index(col_a_lookup) + 1
        except ValueError:
            raise ValueError(f"主表中未找到列: {col_a_lookup}")
        
        # 获取数据范围
        last_row_main = ws_main.range('A' + str(ws_main.cells.last_cell.row)).end('up').row
        if last_row_main < 2:
            last_row_main = ws_main.used_range.last_cell.row
        
        # 确定新列位置
        result_col = self.result_column.get()
        if result_col:  # 如果选择了特定列
            try:
                new_col_idx = header_main.index(result_col) + 1
            except ValueError:
                raise ValueError(f"主表中未找到列: {result_col}")
        else:  # 追加到最后一列
            new_col_idx = len([h for h in header_main if h is not None]) + 1
        
        new_col_name = f"查找结果_{self.column_result_combo.get()}"
        ws_main.cells(1, new_col_idx).value = new_col_name
        
        # 设置新列标题样式
        header_cell = ws_main.cells(1, new_col_idx)
        header_cell.api.Font.Bold = True
        header_cell.api.Interior.Color = 0x4F81BD  # 蓝色背景
        header_cell.api.Font.Color = 0xFFFFFF  # 白色字体
        header_cell.api.HorizontalAlignment = -4108  # 居中
        
        # 单线程处理数据（避免COM对象线程安全问题）
        self.add_message("🔄 正在处理多值查找...")
        
        # 一次性读取所有查找列数据
        lookup_range = ws_main.range(f"{chr(64 + lookup_col_idx)}2:{chr(64 + lookup_col_idx)}{last_row_main}")
        lookup_values = lookup_range.value
        
        # 确保返回的是列表
        if not isinstance(lookup_values, list):
            lookup_values = [lookup_values]
        
        # 单线程处理数据
        matched_count = 0
        not_found_count = 0
        
        # 批量处理，每100行更新一次进度
        batch_size = 100
        total_rows = len(lookup_values)
        
        for i in range(0, total_rows, batch_size):
            end_idx = min(i + batch_size, total_rows)
            
            for j in range(i, end_idx):
                row = j + 2  # 数据从第2行开始
                cell_value = lookup_values[j]
                
                if cell_value is None:
                    ws_main.cells(row, new_col_idx).value = ""
                    continue
                
                # 处理多值查找（按换行符分隔）
                str_value = str(cell_value)
                values = [v.strip() for v in str_value.split('\n') if v.strip()]
                
                if not values:
                    ws_main.cells(row, new_col_idx).value = ""
                    continue
                
                # 对每个值进行查找
                results = []
                for val in values:
                    result = lookup_dict.get(val)
                    if result is None:
                        # 尝试去除空格匹配
                        result = lookup_dict.get(val.strip())
                    
                    if result is not None and result != "":
                        results.append(result)
                        matched_count += 1
                    else:
                        results.append(self.not_found_value)
                        not_found_count += 1
                
                # 用换行符合并结果
                final_result = '\n'.join(results)
                ws_main.cells(row, new_col_idx).value = final_result
            
            # 更新进度
            progress = int((i + batch_size) / total_rows * 50) + 50  # 50-100%范围
            self.update_progress(progress, f"数据处理进度: {min(i + batch_size, total_rows)}/{total_rows} 行")
        
        # 批量设置格式
        self.add_message("🔄 正在设置格式...")
        if last_row_main > 1:
            try:
                # 设置数据单元格格式
                data_range = ws_main.range(f"{chr(64 + new_col_idx)}2:{chr(64 + new_col_idx)}{last_row_main}")
                data_range.api.WrapText = True
                data_range.api.VerticalAlignment = -4108  # 居中
                data_range.api.HorizontalAlignment = -4108  # 居中
                
                # 批量设置行高（优化性能）
                # 使用批量操作替代逐行设置
                if last_row_main - 1 > 1000:  # 大数据量时使用批量设置
                    # 设置整个区域的行高
                    rows_range = ws_main.range(f"2:{last_row_main}")
                    rows_range.api.RowHeight = 50
                else:
                    # 小数据量时逐行设置
                    for row in range(2, last_row_main + 1):
                        ws_main.cells(row, new_col_idx).api.RowHeight = 50
                        
                self.add_message("✅ 格式设置完成")
                
            except Exception as e:
                self.add_message(f"⚠️ 格式设置部分失败，但数据已处理完成: {str(e)}")
                # 继续执行，不中断整个流程
        
        self.add_message(f"✅ 处理完成: 总行数 {total_rows}, 匹配成功 {matched_count}, 未找到 {not_found_count}")
    
    def build_lookup_dict_with_openpyxl(self):
        """使用openpyxl快速构建查找字典"""
        wb_ref = load_workbook(self.file_b_path.get(), data_only=True)
        ws_ref = wb_ref.active
        
        # 获取列索引
        col_b_search = self.column_b_combo.get()
        col_b_result = self.column_result_combo.get()
        
        # 读取表头
        header = [cell.value for cell in ws_ref[1]]
        if not header:
            raise ValueError("参考表没有表头")
        
        try:
            search_col_idx = header.index(col_b_search) + 1
            result_col_idx = header.index(col_b_result) + 1
        except ValueError as e:
            raise ValueError(f"参考表中未找到指定列: {e}")
        
        # 构建查找字典（单线程处理，确保稳定性）
        lookup_dict = {}
        max_row = ws_ref.max_row
        
        # 单线程处理数据
        for row in range(2, max_row + 1):
            search_value = ws_ref.cell(row=row, column=search_col_idx).value
            result_value = ws_ref.cell(row=row, column=result_col_idx).value
            
            if search_value is not None:
                key = str(search_value).strip()
                if result_value is not None:
                    lookup_dict[key] = str(result_value).strip()
                else:
                    lookup_dict[key] = ""
        
        wb_ref.close()
        self.add_message(f"✅ 参考表数据加载完成，共 {len(lookup_dict)} 条记录")
        return lookup_dict
    
    def process_main_table_with_openpyxl(self, lookup_dict):
        """使用openpyxl快速处理主表数据"""
        wb_main = load_workbook(self.file_a_path.get(), data_only=True)
        ws_main = wb_main.active
        
        # 获取列索引
        col_a_lookup = self.column_a_combo.get()
        
        # 读取表头
        header_main = [cell.value for cell in ws_main[1]]
        if not header_main:
            raise ValueError("主表没有表头")
        
        try:
            lookup_col_idx = header_main.index(col_a_lookup) + 1
        except ValueError:
            raise ValueError(f"主表中未找到列: {col_a_lookup}")
        
        # 确定新列位置
        result_col = self.result_column.get()
        if result_col:  # 如果选择了特定列
            try:
                new_col_idx = header_main.index(result_col) + 1
            except ValueError:
                raise ValueError(f"主表中未找到列: {result_col}")
        else:  # 追加到最后一列
            new_col_idx = len([h for h in header_main if h is not None]) + 1
        
        # 添加新列标题
        new_col_name = f"查找结果_{self.column_result_combo.get()}"
        ws_main.cell(row=1, column=new_col_idx).value = new_col_name
        
        # 单线程处理数据（确保稳定性）
        max_row = ws_main.max_row
        total_rows = max_row - 1
        
        self.add_message("🔄 正在处理数据...")
        
        # 单线程处理数据
        for row in range(2, max_row + 1):
            cell_value = ws_main.cell(row=row, column=lookup_col_idx).value
            
            if cell_value is None:
                ws_main.cell(row=row, column=new_col_idx).value = ""
                continue
            
            # 处理多值查找（按换行符分隔）
            str_value = str(cell_value)
            values = [v.strip() for v in str_value.split('\n') if v.strip()]
            
            if not values:
                ws_main.cell(row=row, column=new_col_idx).value = ""
                continue
            
            # 对每个值进行查找
            results = []
            for val in values:
                result = lookup_dict.get(val)
                if result is None:
                    # 尝试去除空格匹配
                    result = lookup_dict.get(val.strip())
                
                if result is not None and result != "":
                    results.append(result)
                else:
                    results.append(self.not_found_value)
            
            # 用换行符合并结果
            final_result = '\n'.join(results)
            ws_main.cell(row=row, column=new_col_idx).value = final_result
            
            # 每100行更新一次进度
            if (row - 1) % 100 == 0:
                progress = int((row - 1) / total_rows * 30) + 40  # 40-70%范围
                self.update_progress(progress, f"数据处理进度: {row - 1}/{total_rows} 行")
        
        # 保存临时文件
        temp_file = f"temp_vlookup_{int(time.time())}.xlsx"
        wb_main.save(temp_file)
        wb_main.close()
        
        self.add_message(f"✅ 数据处理完成，共 {total_rows} 行数据")
        return temp_file
    
    def format_with_xlwings(self, temp_file, output_path):
        """使用xlwings设置格式"""
        app = xw.App(visible=False)
        
        try:
            wb = app.books.open(temp_file)
            ws = wb.sheets[0]
            
            # 获取新列位置
            header_main = ws.range('1:1').value
            new_col_name = f"查找结果_{self.column_result_combo.get()}"
            new_col_idx = header_main.index(new_col_name) + 1
            
            # 设置新列标题样式
            header_cell = ws.cells(1, new_col_idx)
            header_cell.api.Font.Bold = True
            header_cell.api.Interior.Color = 0x4F81BD  # 蓝色背景
            header_cell.api.Font.Color = 0xFFFFFF  # 白色字体
            header_cell.api.HorizontalAlignment = -4108  # 居中
            
            # 设置数据单元格格式
            max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
            if max_row < 2:
                max_row = ws.used_range.last_cell.row
            
            # 批量设置格式
            if max_row > 1:
                try:
                    data_range = ws.range(f"{chr(64 + new_col_idx)}2:{chr(64 + new_col_idx)}{max_row}")
                    data_range.api.WrapText = True
                    data_range.api.VerticalAlignment = -4108  # 居中
                    data_range.api.HorizontalAlignment = -4108  # 居中
                    
                    # 批量设置行高（优化性能）
                    if max_row - 1 > 1000:  # 大数据量时使用批量设置
                        rows_range = ws.range(f"2:{max_row}")
                        rows_range.api.RowHeight = 50
                    else:
                        # 小数据量时逐行设置
                        for row in range(2, max_row + 1):
                            ws.cells(row, new_col_idx).api.RowHeight = 50
                except Exception as e:
                    self.add_message(f"⚠️ 格式设置部分失败，但数据已处理完成: {str(e)}")
                    # 继续执行，不中断整个流程
            
            # 保存最终文件
            wb.save(output_path)
            wb.close()
            
            self.add_message("✅ 格式设置完成")
            
        except Exception as e:
            raise e
        finally:
            app.quit()
    
    def show_completion_message(self):
        """显示完成消息"""
        result_msg = f"""
✅ VLOOKUP操作完成！

📊 结果文件已保存到:
{self.output_file_path}

💡 功能说明：
  • 完美保留原文件所有格式和图片
  • 支持多值查找（换行符分隔）
  • 找不到的值用 {self.not_found_value} 代替
  • 固定行高50磅
  • 内容居中对齐
  • 已保存为新文件，原文件未修改
        """
        
        self.window.after(0, lambda: messagebox.showinfo("完成", result_msg.strip()))
        self.window.after(0, lambda: self.open_result_btn.config(state=tk.NORMAL))
    
    def open_output_file(self):
        """打开输出文件"""
        if self.output_file_path and os.path.exists(self.output_file_path):
            try:
                if sys.platform == 'win32':
                    os.startfile(self.output_file_path)
                elif sys.platform == 'darwin':  # macOS
                    subprocess.call(['open', self.output_file_path])
                else:  # linux
                    subprocess.call(['xdg-open', self.output_file_path])
                self.add_message(f"📂 正在打开结果文件: {os.path.basename(self.output_file_path)}")
            except Exception as e:
                self.add_message(f"❌ 打开文件失败: {str(e)}", is_error=True)
        else:
            self.add_message("❌ 找不到结果文件，请先执行VLOOKUP操作", is_error=True)
    
    def clear_all(self):
        """清空所有选择"""
        self.file_a_path.set("")
        self.file_b_path.set("")
        self.column_a_combo.set("")
        self.column_b_combo.set("")
        self.column_result_combo.set("")
        self.result_column.set("")
        self.output_file_path = None
        self.open_result_btn.config(state=tk.DISABLED)
        
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(tk.END, "✅ 已清空所有选择，请重新选择文件...\n")
        self.status_text.config(state=tk.DISABLED)
    
    def add_message(self, msg, is_error=False):
        """添加消息到队列"""
        self.message_queue.put((msg, is_error))
    
    def process_queue(self):
        """处理消息队列"""
        try:
            while True:
                try:
                    msg, is_error = self.message_queue.get_nowait()
                    self.status_text.config(state=tk.NORMAL)
                    if is_error:
                        self.status_text.insert(tk.END, f"❌ {msg}\n", "error")
                    else:
                        self.status_text.insert(tk.END, f"{msg}\n")
                    self.status_text.config(state=tk.DISABLED)
                    self.status_text.see(tk.END)
                except queue.Empty:
                    break
        finally:
            self.window.after(100, self.process_queue)
    
    def update_progress(self, value, message):
        """更新进度条"""
        def update():
            self.progress["value"] = value
            self.progress_label.config(text=message)
            self.add_message(message)
        self.window.after(0, update)
    
    def return_to_main(self):
        """返回主页"""
        # 检查处理状态
        if self.processing:
            if messagebox.askokcancel("停止处理", "VLOOKUP处理正在进行中，确定要停止并返回主页吗？"):
                if self.xl_app:
                    try:
                        self.xl_app.quit()
                    except:
                        pass
                self.processing = False
                # 调用父级的show_home_page方法
                self.parent.master.show_home_page()
            return
        else:
            # 直接返回主页
            self.parent.master.show_home_page()
    
    def on_close(self):
        """窗口关闭事件处理 - 内嵌模式不需要"""
        # 检查处理状态，但不关闭窗口（内嵌模式由主窗口管理）
        if self.processing:
            if messagebox.askokcancel("停止处理", "VLOOKUP处理正在进行中，确定要停止吗？"):
                if self.xl_app:
                    try:
                        self.xl_app.quit()
                    except:
                        pass
                self.processing = False
        # 内嵌模式下不关闭窗口，由主窗口管理
        # self.window.destroy()
